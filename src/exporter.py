"""
Lead Export utility for CSV, Excel (.xlsx), and JSON formats.
"""
import os
import json
from typing import List, Optional
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.models import BusinessLead


def export_leads(
    leads: List[BusinessLead],
    output_path: str,
    export_format: str = "csv",
) -> str:
    """
    Export leads to specified file format (csv, xlsx, json, or all).
    
    Args:
        leads: List of BusinessLead objects
        output_path: Filepath or base filename for export
        export_format: "csv", "xlsx", "json", or "all"
        
    Returns:
        The primary export filepath
    """
    if not leads:
        return ""

    # Ensure parent directory exists
    dir_name = os.path.dirname(output_path)
    if dir_name:
        os.makedirs(dir_name, exist_ok=True)

    base_path, ext = os.path.splitext(output_path)
    records = [lead.to_flat_dict() for lead in leads]
    df = pd.DataFrame(records)

    target_formats = []
    if export_format == "all":
        target_formats = ["csv", "xlsx", "json"]
    else:
        target_formats = [export_format.lower().lstrip(".")]

    primary_file = ""

    for fmt in target_formats:
        file_to_write = f"{base_path}.{fmt}" if not output_path.endswith(f".{fmt}") else output_path
        if not primary_file:
            primary_file = file_to_write

        if fmt == "csv":
            # Use utf-8-sig so Excel properly renders accents/symbols
            df.to_csv(file_to_write, index=False, encoding="utf-8-sig")

        elif fmt == "xlsx":
            _export_to_styled_excel(df, file_to_write)

        elif fmt in ["json", "jsonl"]:
            if fmt == "jsonl":
                with open(file_to_write, "w", encoding="utf-8") as f:
                    for lead in leads:
                        f.write(json.dumps(lead.model_dump(), ensure_ascii=False) + "\n")
            else:
                with open(file_to_write, "w", encoding="utf-8") as f:
                    json.dump([lead.model_dump() for lead in leads], f, indent=2, ensure_ascii=False)

    return primary_file


def _export_to_styled_excel(df: pd.DataFrame, filepath: str):
    """Write DataFrame to Excel with professional header styling and auto column widths."""
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        sheet_name = "Leads"
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]

        # Style Header
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")

        for col_idx, column in enumerate(worksheet.columns, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

            # Auto-fit column width
            max_len = max(
                len(str(cell.value or "")),
                max((len(str(c.value or "")) for c in column), default=0)
            )
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)
